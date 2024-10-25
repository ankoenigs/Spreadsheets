#!/usr/bin/python3

import pandas as pd
from openpyxl import load_workbook #, Workbook

workbook = load_workbook(filename="test_for_features_and_stuff.xlsx", data_only=False) # not only data so we can extract info such as if it's a formula

print(workbook.sheetnames)

sheet = workbook['Tabelle1']

print(type(sheet))

print(sheet.dimensions)
print(sheet.max_column)
print(sheet.max_row)

# for cell_col in range(1, 4):
#     print(sheet.cell(1, cell_col).data_type) # n - numeric (oder so); f - formula :)
#     print(sheet.cell(1, cell_col).value)
    
is_complex = False

for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        print(sheet.cell(i, j).value)
        if sheet.cell(i, j).data_type == "f":
            is_complex = True
            
print(is_complex)

